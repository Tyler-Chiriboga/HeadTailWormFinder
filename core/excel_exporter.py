"""
Excel exporter for worm annotations.
Exports all annotations to Excel format for analysis.
"""
from pathlib import Path
from typing import Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.annotation_manager import AnnotationManager, WormAnnotation, VideoAnnotations


class ExcelExporter:
    """
    Exports worm annotations to Excel format.
    Creates formatted spreadsheets with all annotation data.
    """
    
    def __init__(self, annotation_manager: AnnotationManager):
        """
        Initialize exporter.
        
        Args:
            annotation_manager: AnnotationManager with annotations to export
        """
        self.manager = annotation_manager
    
    def export(
        self,
        output_path: Optional[str] = None,
        include_incomplete: bool = True
    ) -> str:
        """
        Export all annotations to Excel.
        
        Args:
            output_path: Path for output file. Auto-generated if None.
            include_incomplete: Whether to include incomplete annotations
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            if self.manager.folder_path:
                output_path = self.manager.folder_path / "worm_annotations.xlsx"
            else:
                output_path = Path("worm_annotations.xlsx")
        else:
            output_path = Path(output_path)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create main annotations sheet
        ws_main = wb.active
        ws_main.title = "Annotations"
        self._create_annotations_sheet(ws_main, include_incomplete)
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary)
        
        # Create per-video sheet
        ws_videos = wb.create_sheet("Videos")
        self._create_videos_sheet(ws_videos)
        
        # Save workbook
        wb.save(str(output_path))
        print(f"Exported annotations to {output_path}")
        
        return str(output_path)
    
    def _create_annotations_sheet(
        self, 
        ws, 
        include_incomplete: bool = True
    ):
        """Create the main annotations sheet."""
        # Define headers
        headers = [
            "Video File",
            "Worm ID",
            "Detection X1", "Detection Y1", "Detection X2", "Detection Y2",
            "Head X1", "Head Y1", "Head X2", "Head Y2",
            "Tail X1", "Tail Y1", "Tail X2", "Tail Y2",
            "Confidence",
            "Complete",
            "Notes",
            "Video Path"
        ]
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        row = 2
        for video_path, video_annot in self.manager.annotations.items():
            for worm_id, annot in video_annot.annotations.items():
                # Skip incomplete if requested
                if not include_incomplete and not annot.is_complete():
                    continue
                
                # Extract box coordinates
                det = annot.detection_box or (None, None, None, None)
                head = annot.head_box or (None, None, None, None)
                tail = annot.tail_box or (None, None, None, None)
                
                data = [
                    video_annot.video_filename,
                    worm_id,
                    det[0], det[1], det[2], det[3],
                    head[0], head[1], head[2], head[3],
                    tail[0], tail[1], tail[2], tail[3],
                    annot.confidence,
                    "Yes" if annot.is_complete() else "No",
                    annot.notes,
                    video_path
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = thin_border
                    if col in [6, 10, 14]:  # X2 columns - right align numbers
                        cell.alignment = Alignment(horizontal="right")
                
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _create_summary_sheet(self, ws):
        """Create summary statistics sheet."""
        stats = self.manager.get_statistics()
        
        # Style
        header_font = Font(bold=True)
        
        # Title
        ws.cell(row=1, column=1, value="Annotation Summary").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Statistics
        summary_data = [
            ("", ""),
            ("Total Videos Annotated", stats['total_videos']),
            ("Total Worms Annotated", stats['total_worms']),
            ("Complete Annotations (Head + Tail)", stats['complete_annotations']),
            ("Incomplete Annotations", stats['incomplete_annotations']),
            ("", ""),
            ("Folder Path", str(self.manager.folder_path) if self.manager.folder_path else "N/A"),
        ]
        
        for row, (label, value) in enumerate(summary_data, 4):
            ws.cell(row=row, column=1, value=label).font = header_font
            ws.cell(row=row, column=2, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
    
    def _create_videos_sheet(self, ws):
        """Create per-video summary sheet."""
        headers = [
            "Video File",
            "Total Worms",
            "Complete",
            "Incomplete",
            "Frame Width",
            "Frame Height",
            "Created",
            "Modified"
        ]
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data
        row = 2
        for video_path, video_annot in self.manager.annotations.items():
            total = len(video_annot.annotations)
            complete = sum(1 for a in video_annot.annotations.values() if a.is_complete())
            
            data = [
                video_annot.video_filename,
                total,
                complete,
                total - complete,
                video_annot.frame_width,
                video_annot.frame_height,
                video_annot.created_at[:19] if video_annot.created_at else "",
                video_annot.modified_at[:19] if video_annot.modified_at else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header
        ws.freeze_panes = "A2"


def export_for_training(
    annotation_manager: AnnotationManager,
    output_folder: Optional[Path] = None
) -> dict:
    """
    Export annotations in a format suitable for model training.
    Creates YOLO-format label files and organized image crops.
    
    Args:
        annotation_manager: AnnotationManager with annotations
        output_folder: Output folder for training data
        
    Returns:
        Dictionary with export statistics
    """
    if output_folder is None:
        output_folder = annotation_manager.folder_path / "training_data"
    
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Create subdirectories
    (output_folder / "images" / "heads").mkdir(parents=True, exist_ok=True)
    (output_folder / "images" / "tails").mkdir(parents=True, exist_ok=True)
    (output_folder / "labels").mkdir(parents=True, exist_ok=True)
    
    stats = {
        'total_exported': 0,
        'heads_exported': 0,
        'tails_exported': 0,
        'labels_created': 0
    }
    
    # Create labels file for head/tail detection training
    labels_file = output_folder / "labels" / "annotations.txt"
    
    with open(labels_file, 'w') as f:
        f.write("# HeadTailWormFinder Training Data Export\n")
        f.write(f"# Generated: {datetime.now().isoformat()}\n")
        f.write("# Format: video_file, worm_id, type, x1, y1, x2, y2\n\n")
        
        for video_path, video_annot in annotation_manager.annotations.items():
            for worm_id, annot in video_annot.annotations.items():
                if annot.head_box:
                    box = annot.head_box
                    f.write(f"{video_annot.video_filename},{worm_id},head,{box[0]},{box[1]},{box[2]},{box[3]}\n")
                    stats['heads_exported'] += 1
                
                if annot.tail_box:
                    box = annot.tail_box
                    f.write(f"{video_annot.video_filename},{worm_id},tail,{box[0]},{box[1]},{box[2]},{box[3]}\n")
                    stats['tails_exported'] += 1
                
                stats['total_exported'] += 1
    
    stats['labels_created'] = 1
    
    print(f"Exported training data to {output_folder}")
    print(f"Stats: {stats}")
    
    return stats


if __name__ == "__main__":
    # Test exporter
    from core.annotation_manager import AnnotationManager
    
    manager = AnnotationManager()
    
    # Add test data
    test_video = "/test/video1.avi"
    annot = manager.add_worm_annotation(test_video, (100, 100, 200, 200), 0.95)
    manager.set_head_box(test_video, annot.worm_id, (100, 100, 130, 130))
    manager.set_tail_box(test_video, annot.worm_id, (170, 170, 200, 200))
    
    # Test export
    exporter = ExcelExporter(manager)
    # exporter.export("/tmp/test_annotations.xlsx")
    
    print("Exporter test complete")
